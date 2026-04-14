from sqlalchemy.orm import Session, joinedload, aliased
from fastapi import HTTPException
from datetime import date, datetime
from typing import Optional
from sqlalchemy import extract, func
from math import ceil
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

from app.models.overtime_request import OvertimeRequest
from app.models.absence import ApprovalStatus
from app.schemas.overtime_request import OvertimeCreate, OvertimeApprove
from app.models.employee import Employee
from app.models.shift import Shift
from app.core.config import OVER_TIME_MONTHLY_LIMIT_MINS, OVER_TIME_YEARLY_LIMIT_MINS, OVERTIME_DAILY_LIMIT_MINS
from app.models.attendance_log import AttendanceLog
from app.services.calendar_service import calendar_service
from app.models.daily_work_report import DailyWorkReport

class OvertimeService:
    def create_request(self, db: Session, obj_in: OvertimeCreate, emp_id: int):
        multiplier_value = obj_in.ot_type.multiplier
        
        employee = db.query(Employee).options(joinedload(Employee.shift)).filter(
            Employee.id == emp_id
        ).first()
        if not employee or not employee.shift:
            raise HTTPException(
                status_code=404, 
                detail="Không tìm thấy thông tin nhân viên hoặc ca làm việc của nhân viên này."
            )

        shift : Shift = employee.shift
        
        is_outside_shift = (obj_in.end_time <= shift.start_time) or (obj_in.start_time >= shift.end_time)

        if not is_outside_shift:
            raise HTTPException(
                status_code=400,
                detail=f"Thời gian OT không được trùng với ca làm việc {shift.name}."
            )

        current_month_total = db.query(func.sum(OvertimeRequest.actual_work_time)).filter(
            OvertimeRequest.employee_id == emp_id,
            OvertimeRequest.status != ApprovalStatus.REJECTED,
            extract('month', OvertimeRequest.work_date) == obj_in.work_date.month,
            extract('year', OvertimeRequest.work_date) == obj_in.work_date.year
        ).scalar() or 0

        calc_minutes = (obj_in.end_time.hour * 60 + obj_in.end_time.minute) - \
                (obj_in.start_time.hour * 60 + obj_in.start_time.minute)

        if calc_minutes > OVERTIME_DAILY_LIMIT_MINS:
            raise HTTPException(
                status_code=400,
                detail=f"Vi phạm định mức: Thời gian OT không được vượt quá {OVERTIME_DAILY_LIMIT_MINS/60} giờ một ngày"
            )

        if current_month_total + calc_minutes > OVER_TIME_MONTHLY_LIMIT_MINS:
            raise HTTPException(
                status_code=400,
                detail=f"Vi phạm định mức: Tổng giờ làm thêm trong tháng không được quá 40 giờ. "
                    f"Bạn đã làm {current_month_total/60:.1f}h, đơn này thêm {calc_minutes/60:.1f}h."
            )

        current_year_total = db.query(func.sum(OvertimeRequest.actual_work_time)).filter(
            OvertimeRequest.employee_id == emp_id,
            OvertimeRequest.status != ApprovalStatus.REJECTED,
            extract('year', OvertimeRequest.work_date) == obj_in.work_date.year
        ).scalar() or 0

        if current_year_total + calc_minutes > OVER_TIME_YEARLY_LIMIT_MINS:
            raise HTTPException(
                status_code=400,
                detail=f"Vi phạm định mức: Tổng giờ làm thêm trong năm không được quá 200 giờ. "
                    f"Hiện tại đã tích lũy {current_year_total/60:.1f}h."
            )

        db_obj = OvertimeRequest(
            **obj_in.model_dump(),
            employee_id=emp_id,
            multiplier=multiplier_value,
            status=ApprovalStatus.PENDING
        )
        db.add(db_obj)
        db.commit()
        db.refresh(db_obj)
        return db_obj

    def get_my_requests(
        self, 
        db: Session, 
        emp_id: int, 
        month: Optional[int] = None, 
        year: Optional[int] = None
    ):
        query = db.query(OvertimeRequest).filter(OvertimeRequest.employee_id == emp_id)
        
        if month:
            query = query.filter(extract('month', OvertimeRequest.work_date) == month)
        if year:
            query = query.filter(extract('year', OvertimeRequest.work_date) == year)
            
        return query.order_by(OvertimeRequest.work_date.desc()).all()
    
    def get_all_requests_admin(self, db: Session, month: int, year: int, page: int, limit: int, status: str = None, search: str = None):
        query = db.query(OvertimeRequest).join(
                Employee, OvertimeRequest.employee_id == Employee.id
            ).add_columns(Employee.full_name.label("employee_name"))
        
        query = query.filter(extract('month', OvertimeRequest.work_date) == month)
        query = query.filter(extract('year', OvertimeRequest.work_date) == year)

        if status:
            query = query.filter(OvertimeRequest.status == status)

        if search:
            query = query.filter(Employee.id == search)

        total_elements = query.count()
        offset = (page - 1) * limit
        raw_results = query.order_by(
            OvertimeRequest.status == 'PENDING', 
            OvertimeRequest.created_at.desc()
            ).offset(offset).limit(limit).all()

        final_results = []
        for row in raw_results:

            obj = row[0] 
            emp_name = row[1] 
            
            obj.employee_name = emp_name 
            final_results.append(obj)

        total_pages = ceil(total_elements / limit) if total_elements > 0 else 0
            
        return final_results, {
            "total_elements": total_elements, 
            "total_pages": total_pages, 
            "page": page, 
            "limit": limit
        }
    
    def delete_pending_request(self, db: Session, ot_id: int, emp_id: int):
        db_obj = db.query(OvertimeRequest).filter(
            OvertimeRequest.id == ot_id,
            OvertimeRequest.employee_id == emp_id
        ).first()

        if not db_obj:
            raise HTTPException(404, "Không tìm thấy đơn OT")

        if db_obj.status != ApprovalStatus.PENDING:
            raise HTTPException(400, "Chỉ có thể xóa đơn đang chờ duyệt (PENDING)")

        db.delete(db_obj)
        db.commit()
        return True

    def approve_request(self, db: Session, ot_id: int, admin_id: int, obj_in: OvertimeApprove):
        db_obj = db.query(OvertimeRequest).filter(OvertimeRequest.id == ot_id).first()
        if not db_obj:
            raise HTTPException(404, "Không tìm thấy đơn OT")

        db_obj.status = obj_in.status
        db_obj.approved_by = admin_id
        db_obj.approved_at = datetime.now()
        
        db.commit()
        db.refresh(db_obj)
        return db_obj
    
    def _time_to_minutes(self, t) -> int:
        if t is None: return 0
        return t.hour * 60 + t.minute

    def calculate_actual_ot_minutes(self, db: Session, ot_id: int):
        ot_req = db.query(OvertimeRequest).filter(OvertimeRequest.id == ot_id).first()
        if not ot_req:
            return None

        working_days = calendar_service.get_working_days_list(
            db, ot_req.work_date, ot_req.work_date
        )
        is_working_day = len(working_days) > 0

        registered_min = (
            self._time_to_minutes(ot_req.end_time)
            - self._time_to_minutes(ot_req.start_time)
        )

        actual_ot = 0

        if is_working_day:
            # Nếu hôm đó làm việc, lấy ot từ DailyWorkReport
            report = db.query(DailyWorkReport).filter(
                DailyWorkReport.employee_id == ot_req.employee_id,
                DailyWorkReport.work_date == ot_req.work_date
            ).first()

            if report:
                actual_ot = report.overtime_minutes or 0
            else:
                actual_ot = 0

        else:
            # nếu hôm đó không làm việc, tính từ log điểm danh
            logs = db.query(AttendanceLog).filter(
                AttendanceLog.employee_id == ot_req.employee_id,
                AttendanceLog.log_date == ot_req.work_date
            ).order_by(AttendanceLog.checked_time.asc()).all()

            if not logs:
                ot_req.actual_work_time = 0
                db.commit()
                return 0

            first_log_min = self._time_to_minutes(logs[0].checked_time)
            last_log_min = self._time_to_minutes(logs[-1].checked_time)

            total_presence_min = last_log_min - first_log_min

            lunch_break = 60
            actual_ot = total_presence_min - lunch_break

        final_ot_min = max(0, min(actual_ot, registered_min))

        ot_req.actual_work_time = final_ot_min
        db.commit()

        return final_ot_min

    def batch_process_actual_ot(self, db: Session, target_date: date):
        """
        Hàm chạy định kỳ để cập nhật giờ OT thực tế cho tất cả đơn trong 1 ngày
        """
        requests = db.query(OvertimeRequest).filter(
            OvertimeRequest.work_date == target_date,
            OvertimeRequest.status == ApprovalStatus.APPROVED
        ).all()
        
        for req in requests:
            self.calculate_actual_ot_minutes(db, req.id)
        
        return len(requests)
    
    def export_overtime_excel(self, db: Session, month: int = None, year: int = None):
        # Tạo alias để join bảng nhân viên 2 lần
        Requester = aliased(Employee)
        Approver = aliased(Employee)

        query = db.query(
            OvertimeRequest,
            Requester.full_name.label("employee_name"),
            Approver.full_name.label("approver_name")
        ).join(Requester, OvertimeRequest.employee_id == Requester.id)\
         .outerjoin(Approver, OvertimeRequest.approved_by == Approver.id) # Dùng outerjoin vì đơn PENDING chưa có người duyệt

        if month and year:
            query = query.filter(
                extract('month', OvertimeRequest.work_date) == month,
                extract('year', OvertimeRequest.work_date) == year
            )

        results = query.all()

        data_list = []
        for ot, emp_name, admin_name in results:
            data_list.append({
                "Ngày làm": ot.work_date,
                "Mã NV": ot.employee_id,
                "Tên nhân viên": emp_name,
                "Bắt đầu": ot.start_time.strftime("%H:%M") if ot.start_time else "",
                "Kết thúc": ot.end_time.strftime("%H:%M") if ot.end_time else "",
                "Công thực tế (phút)": ot.actual_work_time or 0,
                "Loại OT": ot.ot_type.value if hasattr(ot.ot_type, 'value') else ot.ot_type,
                "Hệ số": ot.multiplier,
                "Trạng thái": ot.status.value if hasattr(ot.status, 'value') else ot.status,
                "Lý do": ot.reason,
                "Người duyệt": admin_name or "---"
            })

        df = pd.DataFrame(data_list)
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = 'Danh sách OT'
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # --- Apply Styles ---
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")

            # Style cho Header
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # Auto-width và Center Align cho dữ liệu
            for col in worksheet.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)

                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                    
                    # Căn giữa số và dữ liệu chung
                    cell.alignment = center_align
                
                worksheet.column_dimensions[col_letter].width = max_length + 4

            # Style đặc biệt: Highlight màu đỏ nếu trạng thái là REJECTED
            status_col_index = list(df.columns).index("Trạng thái") + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=status_col_index)
                if cell.value == "REJECTED" or cell.value == "rejected":
                    cell.font = Font(color="FF0000", bold=True)
                elif cell.value == "APPROVED" or cell.value == "approved":
                    cell.font = Font(color="28a745", bold=True)

        output.seek(0)
        file_name = f"Bao_cao_OT_{month}_{year}.xlsx"
        return output, file_name

overtime_service = OvertimeService()