import calendar
from datetime import date, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import func
from fastapi import HTTPException
from datetime import datetime
from typing import List, Dict
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.models.employee import Employee
from app.models.daily_work_report import DailyWorkReport
from app.models.monthly_report import MonthlyWorkReport
from app.models.absence import AbsenceType, ApprovalStatus, Absence
from app.models.attendance_correction import AttendanceCorrection
from app.models.timesheet_period_control import TimesheetPeriodControl
from app.models.absence_plan import AbsencePlan
from app.services.calendar_service import calendar_service

# lưu ý: công chuẩn là 8 tiếng = 480 phút. Nếu ngày công là 500 phút, chỉ lấy 480 phút
# Tính OT: sẽ để OT request xử lý vì vướng logic hệ số lương
class PayrollService:
    def _get_debt_adjustment(self, db: Session, employee_id: int, first_day_current_month: date):
        last_month_end_date = first_day_current_month - timedelta(days=1)
        first_day_last_month = last_month_end_date.replace(day=1)

        # lấy thông tin kỳ chốt trước
        last_period_control = db.query(TimesheetPeriodControl).filter(
            TimesheetPeriodControl.month == first_day_last_month.month,
            TimesheetPeriodControl.year == first_day_last_month.year
        ).first()

        if not last_period_control:
            return 0

        # lấy báo cáo tháng trước
        last_month_report = db.query(MonthlyWorkReport).filter(
            MonthlyWorkReport.employee_id == employee_id,
            MonthlyWorkReport.period_start == first_day_last_month
        ).first()

        if not last_month_report or last_month_report.estimated_minutes == 0:
            return 0

        # tính nợ công = Công chuẩn thực tế + Công nghỉ phép có lương - Công tạm tính tháng trước
        start_check = last_period_control.closing_date
        end_check = last_month_end_date

        actual_reports = db.query(DailyWorkReport).filter(
            DailyWorkReport.employee_id == employee_id,
            DailyWorkReport.work_date >= start_check,
            DailyWorkReport.work_date <= end_check
        ).all()
        
        corrected_dates = {
            c.work_date for c in db.query(AttendanceCorrection.work_date).filter(
                AttendanceCorrection.employee_id == employee_id,
                AttendanceCorrection.work_date >= start_check,
                AttendanceCorrection.work_date <= end_check,
                AttendanceCorrection.status == ApprovalStatus.APPROVED
            ).all()
        }

        standard_work_minutes = 0
        #  nếu được fix công
        for report in actual_reports:
            if report.work_date in corrected_dates:
                standard_work_minutes += 480
            else:
                standard_work_minutes += min(report.work_time_minutes or 0, 480)

        paid_leave_count = db.query(func.count(Absence.id)).filter(
            Absence.employee_id == employee_id,
            Absence.work_date >= start_check,
            Absence.work_date <= end_check,
            Absence.is_paid == True
        ).scalar() or 0

        paid_minutes = paid_leave_count * 480

        debt = standard_work_minutes + paid_minutes - last_month_report.estimated_minutes
        return debt

    def _calculate_estimated_minutes(self, db: Session, start_date: date, end_date: date):
        """
        Tính phút tạm tính dựa trên danh sách ngày làm việc thực tế 
        (đã loại trừ T7, CN và các ngày Lễ trong bảng Vacation)
        """
        working_days = calendar_service.get_working_days_list(db, start_date, end_date)

        return len(working_days) * 480
    
    def _calculate_actual_metrics(self, db: Session, employee_id: int, start_date: date, end_date: date) -> Dict:
        metrics = {
            "standard_work_minutes": 0,
            "lack_minutes": 0,
            "actual_work_days": 0,
            "paid_leave_days": 0,
            "unpaid_leave_days": 0
        }

        actual_reports = db.query(DailyWorkReport).filter(
            DailyWorkReport.employee_id == employee_id,
            DailyWorkReport.work_date >= start_date,
            DailyWorkReport.work_date < end_date
        ).all()

        for report in actual_reports:
            # Check đơn sửa công
            has_correction = db.query(AttendanceCorrection).filter(
                AttendanceCorrection.employee_id == employee_id,
                AttendanceCorrection.work_date == report.work_date,
                AttendanceCorrection.status == ApprovalStatus.APPROVED
            ).first()

            if has_correction:
                metrics["standard_work_minutes"] += 480
                metrics["actual_work_days"] += 1
                continue

            # Lấy công chuẩn (max 480p)
            daily_standard = min(report.work_time_minutes or 0, 480)
            metrics["standard_work_minutes"] += daily_standard
            metrics["lack_minutes"] += (report.lack_minutes or 0)

            if daily_standard > 0:
                metrics["actual_work_days"] += 1

        # 2. Xử lý nghỉ phép (Absence)
        absences = db.query(Absence).filter(
            Absence.employee_id == employee_id,
            Absence.work_date >= start_date,
            Absence.work_date < end_date, # < ngày chốt công
        ).all()

        for abs_obj in absences:
            if abs_obj.is_paid:
                metrics["paid_leave_days"] += 1
            else:
                metrics["unpaid_leave_days"] += 1

        return metrics
    
    def close_monthly_payroll(self, db: Session, employee_id: int, closing_day: int, month: int, year: int):
        closing_date = date(year, month, closing_day)

        first_day = closing_date.replace(day=1) 
        last_day_of_month = closing_date.replace(day=calendar.monthrange(closing_date.year, closing_date.month)[1])

        # Thai sản sẽ xử lý riêng
        is_special_leave = db.query(AbsencePlan).filter(
            AbsencePlan.employee_id == employee_id,
            AbsencePlan.status == ApprovalStatus.APPROVED,
            AbsencePlan.start_date <= last_day_of_month,
            AbsencePlan.end_date >= first_day,
            AbsencePlan.absence_type == AbsenceType.MATERNITY
        ).first()

        if is_special_leave:
            print(f"Bỏ qua NV {employee_id} do đang nghỉ chế độ.")
            return None

        actual_metrics = self._calculate_actual_metrics(db, employee_id, first_day, closing_date)

        # trừ đi khoản nợ công tháng trước
        debt_adjustment = self._get_debt_adjustment(db, employee_id, first_day)
        adjusted_total = actual_metrics["standard_work_minutes"] + debt_adjustment

        # tạm tính những ngày còn lại
        estimated_minutes = self._calculate_estimated_minutes(db, closing_date, last_day_of_month)

        monthly_log = MonthlyWorkReport(
            employee_id=employee_id,
            period_start=first_day,
            period_end=closing_date,
            standard_work_minutes=adjusted_total,
            lack_minutes=actual_metrics["lack_minutes"],
            estimated_minutes=estimated_minutes,
            actual_work_days=actual_metrics["actual_work_days"],
            paid_leave_days=actual_metrics["paid_leave_days"],
            unpaid_leave_days=actual_metrics["unpaid_leave_days"]
        )

        db.add(monthly_log)
        return monthly_log

    def calculate_all_employees_payroll(self, db: Session, closing_day: int, month: int, year: int) -> Dict:
        """ Quản lý Transaction. """
        
        # kỳ công này đã bị khóa chưa
        period_control = db.query(TimesheetPeriodControl).filter(
            TimesheetPeriodControl.month == month,
            TimesheetPeriodControl.year == year
        ).first()

        if period_control and period_control.is_locked:
            raise HTTPException(
                status_code=400, 
                detail=f"Kỳ công tháng {month}/{year} đã bị khóa. Không thể tính toán lại."
            )

        try:
            # 2. Xóa hết dữ liệu cũ trong MonthlyWorkReport của tháng này để tính lại từ đầu
            first_day_of_month = date(year, month, 1)
            
            db.query(MonthlyWorkReport).filter(
                MonthlyWorkReport.period_start == first_day_of_month
            ).delete(synchronize_session=False)
            
            db.flush()

            employees: List[Employee] = db.query(Employee).filter(Employee.is_active == True).all()
            processed_count = 0
            skipped_count = 0

            for emp in employees:
                result = self.close_monthly_payroll(
                    db=db, 
                    employee_id=emp.id, 
                    closing_day=closing_day, 
                    year=year, 
                    month=month
                )
                if result:
                    processed_count += 1
                else:
                    skipped_count += 1

            db.commit()
            
            return {
                "status": "success",
                "period": f"{month}/{year}",
                "total_employees": len(employees),
                "processed": processed_count,
                "skipped_special_leave": skipped_count,
                "message": "Dữ liệu cũ đã được làm mới và tính toán lại thành công."
            }

        except Exception as e:
            db.rollback()
            print(f"Transaction failed: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Quá trình tính toán thất bại: {str(e)}")

    def lock_timesheet_period(self, db: Session, admin_id: int, month: int, year: int, closing_day: int = 20):
        """
        Sau khi thực hiện chốt công cho toàn bộ nhân viên và khóa kỳ công lại, sau khi khóa không cho fix request nữa
        """
        existing_period = db.query(TimesheetPeriodControl).filter(
            TimesheetPeriodControl.month == month,
            TimesheetPeriodControl.year == year
        ).first()

        if existing_period and existing_period.is_locked:
            raise HTTPException(status_code=400, detail=f"Kỳ công tháng {month}/{year} đã được khóa trước đó.")

        if not existing_period:
            new_control = TimesheetPeriodControl(
                month=month,
                year=year,
                closing_date=date(year, month, closing_day),
                is_locked=True,
                locked_by=admin_id,
                locked_at=datetime.now(),
                note=f"Chốt công tháng {month} năm {year} vào ngày {closing_day}"
            )
            db.add(new_control)
        else:
            existing_period.is_locked = True
            existing_period.locked_by = admin_id
            existing_period.locked_at = datetime.now()
            existing_period.closing_date = date(year, month, closing_day)

        db.commit()
        return existing_period or new_control
    
    def get_monthly_reports(self, db: Session, page: int = 1, limit: int = 30, employee_name: str = None):
        query = db.query(
            MonthlyWorkReport,
            Employee.full_name.label("employee_name"),
            Employee.department_id,
            Employee.email
        ).join(Employee, MonthlyWorkReport.employee_id == Employee.id)

        if employee_name:
            query = query.filter(Employee.full_name.ilike(f"%{employee_name}%"))

        total_elements = query.count()
        offset = (page - 1) * limit
        results = query.offset(offset).limit(limit).all()

        final_data = []
        for row in results:
            report = row[0]  # Lấy object MonthlyWorkReport
            
            report.employee_name = row[1]
            report.department_id = row[2]
            report.email = row[3]
            
            final_data.append(report)

        return {
            "data": final_data,
            "pagination": {
                "total_elements": total_elements,
                "total_pages": (total_elements + limit - 1) // limit,
                "page": page,
                "limit": limit
            }
        }
        
    def get_timesheet_period(self, db: Session, month: int, year: int):
        """
        Lấy thông tin kỳ chốt công dựa trên tháng và năm
        """
        period = db.query(TimesheetPeriodControl).filter(
            TimesheetPeriodControl.month == month,
            TimesheetPeriodControl.year == year
        ).first()
        
        return period
    
    def export_payroll_excel(self, db: Session, month: int = None, year: int = None):

        query = db.query(
            MonthlyWorkReport,
            Employee.full_name,
            Employee.email,
            Employee.department_id
        ).join(Employee, MonthlyWorkReport.employee_id == Employee.id)

        if month and year:
            from sqlalchemy import extract
            query = query.filter(
                extract('month', MonthlyWorkReport.period_start) == month,
                extract('year', MonthlyWorkReport.period_start) == year
            )

        results = query.all()

        data_list = []
        for report, name, email, dept_id in results:
            data_list.append({
                "ID": report.id,
                "Mã NV": report.employee_id,
                "Tên nhân viên": name,
                "Email": email,
                "Phòng ban": dept_id,
                "Kỳ công": f"{report.period_start} - {report.period_end}",
                "Công thực tế (giờ)": round(report.standard_work_minutes / 60, 2),
                "Nợ (phút)": report.lack_minutes,
                "Công tạm tính (giờ)": round(report.estimated_minutes / 60, 2),
                "Ngày làm": report.actual_work_days,
                "Nghỉ có lương": report.paid_leave_days,
                "Nghỉ không lương": report.unpaid_leave_days
            })

        df = pd.DataFrame(data_list)

        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Bảng công')

            workbook = writer.book
            worksheet = writer.sheets['Bảng công']

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")

            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            for col in worksheet.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)

                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                    # căn giữa số
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = center_align

                worksheet.column_dimensions[col_letter].width = max_length + 2

            debt_col_index = list(df.columns).index("Nợ (phút)") + 1

            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=debt_col_index)
                if cell.value and cell.value > 0:
                    cell.font = Font(color="FF0000", bold=True)

        output.seek(0)

        file_name = f"Bang_cong_{month}-{year}.xlsx" if month and year else "Bang_cong.xlsx"

        return output, file_name

payroll_service = PayrollService()